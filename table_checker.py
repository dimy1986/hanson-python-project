# -*- coding: utf-8 -*-
"""
表格检查工具 v2 - 针对复杂非结构化表格
功能：
1. 循环读取 d:/待检查文件夹 下的所有Excel表格
2. 根据三个规则进行检查
3. 输出结果到 待检查文件夹/检查结果.xlsx

表格结构：
- 行1-5: 基本信息（员工号、所属机构、岗位等）
- 行5: "2025年度是否需要参加测评" 字段值
- 行6: "持有合格证有效期" 字段值
- 行9: "第三方测评培训时间" 字段值
- 下方: 考试成绩表（多行，每行一条成绩）
"""

import os
import pandas as pd
from datetime import datetime
import traceback
from openpyxl import load_workbook


def parse_date_range(date_str):
    """
    解析日期范围，如 '2023.8-2026.8' 或 '2023/8-2026/8'
    返回 (start_year, start_month, end_year, end_month)
    """
    if not date_str or pd.isna(date_str):
        return None
    
    try:
        date_str = str(date_str).strip()
        # 替换不同的分隔符
        date_str = date_str.replace('/', '.').replace('～', '-').replace('~', '-')
        
        # 分割开始和结束日期
        parts = date_str.split('-')
        if len(parts) != 2:
            return None
        
        start_part = parts[0].strip()
        end_part = parts[1].strip()
        
        # 解析开始日期
        start_dates = start_part.split('.')
        if len(start_dates) != 2:
            return None
        start_year = int(start_dates[0])
        start_month = int(start_dates[1])
        
        # 解析结束日期
        end_dates = end_part.split('.')
        if len(end_dates) != 2:
            return None
        end_year = int(end_dates[0])
        end_month = int(end_dates[1])
        
        return (start_year, start_month, end_year, end_month)
    except Exception as e:
        print(f"      ❌ 日期解析错误: {date_str}, 错误: {e}")
        return None


def is_year_in_range(year, date_range):
    """
    判断指定年份是否在有效期范围内
    """
    if not date_range:
        return False
    
    start_year, start_month, end_year, end_month = date_range
    return start_year <= year <= end_year


def read_excel_cell_value(file_path, row, col):
    """
    直接读取Excel特定单元格的值（用于处理合并单元格）
    row, col: 1-based索引
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        cell = ws.cell(row=row, column=col)
        return cell.value
    except Exception as e:
        print(f"      ❌ 读取单元格出错: {e}")
        return None


def check_single_file(file_path):
    """
    检查单个Excel文件 - 针对复杂非结构化表格
    """
    file_errors = []
    file_name = os.path.basename(file_path)
    
    print(f"\n📄 检查文件: {file_name}")
    
    try:
        # 使用openpyxl读取原始内容以处理合并单元格
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"   📋 工作表名: {ws.title}")
        print(f"   📊 行数: {ws.max_row}, 列数: {ws.max_column}\n")
        
        # ===== 规则1: 2025年度是否需要参加测评 =====
        print(f"   [规则1] 检查2025年度是否需要参加测评")
        
        # 查找"2025年度是否需要参加测评"和"持有合格证有效期"的位置
        # 通常在第5-6行附近
        test_2025_value = None
        valid_period_value = None
        test_2025_row = None
        valid_period_row = None
        
        for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell_value = str(cell.value).strip() if cell.value else ""
                
                # 查找"2025年度是否需要参加测评"标签及其对应的值
                if "2025年度是否需要参加测评" in cell_value:
                    # 该行的值通常在同一行或下一行的右边单元格
                    test_2025_row = cell.row
                    # 查找该行的黄色高亮单元格（通常是值）
                    for c in ws[cell.row]:
                        if c.value and str(c.value).strip() in ['是', '否']:
                            test_2025_value = str(c.value).strip()
                            break
                    print(f"      找到'2025年度是否需要参加测评' at 行{cell.row}, 值: {test_2025_value}")
                
                # 查找"持有合格证有效期"标签及其对应的值
                if "持有合格证有效期" in cell_value:
                    valid_period_row = cell.row
                    # 查找该行右边的值（通常是日期范围）
                    for c in ws[cell.row]:
                        val = str(c.value).strip() if c.value else ""
                        if '-' in val and ('.' in val or '/' in val):
                            valid_period_value = val
                            break
                    print(f"      找到'持有合格证有效期' at 行{cell.row}, 值: {valid_period_value}")
        
        if test_2025_value and valid_period_value:
            # 解析有效期
            date_range = parse_date_range(valid_period_value)
            if date_range:
                is_in_range = is_year_in_range(2025, date_range)
                expected_value = "否" if is_in_range else "是"
                
                print(f"      有效期: {valid_period_value} -> {date_range}")
                print(f"      2025在范围内: {is_in_range}")
                print(f"      期望值: {expected_value}, 实际值: {test_2025_value}")
                
                if test_2025_value != expected_value:
                    print(f"      ❌ 检查失败")
                    file_errors.append({
                        'file': file_name,
                        'row': test_2025_row,
                        'check_type': '规则1',
                        'field': '2025年度是否需要参加测评',
                        'expected': expected_value,
                        'actual': test_2025_value,
                        'message': f"2025年度是否需要参加测评值不对，应为'{expected_value}'，实际为'{test_2025_value}'"
                    })
                else:
                    print(f"      ✅ 检查通过")
            else:
                print(f"      ⚠️ 无法解析有效期: {valid_period_value}")
        else:
            print(f"      ⚠️ 未找到必要字段")
        
        # ===== 规则2: 第三方测评培训时间 =====
        print(f"\n   [规则2] 检查第三方测评培训时间")
        
        training_time_value = None
        
        for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell_value = str(cell.value).strip() if cell.value else ""
                
                if "第三方测评培训时间" in cell_value or "培训时间" in cell_value:
                    # 查找该行右边的日期值
                    for c in ws[cell.row]:
                        val = str(c.value).strip() if c.value else ""
                        if val and ('/' in val or '.' in val) and len(val) >= 6:
                            training_time_value = val
                            break
                    print(f"      找到培训时间 at 行{cell.row}, 值: {training_time_value}")
        
        if training_time_value and valid_period_value:
            date_range = parse_date_range(valid_period_value)
            if date_range:
                start_year, start_month, end_year, end_month = date_range
                
                # 解析培训时间
                training_parts = training_time_value.replace('/', '.').split('.')
                try:
                    if len(training_parts) >= 2:
                        training_year = int(training_parts[0])
                        training_month = int(training_parts[1])
                        
                        print(f"      有效期开始: {start_year}.{start_month}")
                        print(f"      培训时间: {training_year}.{training_month}")
                        
                        if training_year == start_year and training_month < start_month:
                            print(f"      ✅ 检查通过")
                        else:
                            print(f"      ❌ 检查失败")
                            file_errors.append({
                                'file': file_name,
                                'row': '',
                                'check_type': '规则2',
                                'field': '第三方测评培训时间',
                                'expected': f'在{start_year}.{start_month}之前的同一年',
                                'actual': training_time_value,
                                'message': f"第三方测评培训时间不对，应在有效期开始日期({start_year}.{start_month})之前的同一年"
                            })
                except:
                    print(f"      ⚠️ 无法解析培训时间格式")
            else:
                print(f"      ⚠️ 无法解析有效期")
        else:
            print(f"      ⚠️ 未找到培训时间或有效期")
        
        # ===== 规则3: 考试成绩 =====
        print(f"\n   [规则3] 检查考试成绩")
        
        low_score_rows = []
        
        # 查找考试成绩表（通常在中间偏下的位置）
        for row_idx, row in enumerate(ws.iter_rows(min_row=10, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=10):
            row_data = [str(cell.value).strip() if cell.value else "" for cell in row]
            row_str = "".join(row_data)
            
            # 查找包含成绩的行（通常有"考试成绩"或数字）
            for cell_idx, cell in enumerate(row):
                if cell.value:
                    try:
                        # 尝试将单元格转换为数字
                        score_value = float(cell.value)
                        
                        # 检查是否在考试成绩表的范围内（通过上下文判断）
                        # 如果这一行包含日期或"考试"相关词汇
                        row_context = "".join(str(c.value) for c in row if c.value)
                        if any(x in row_context for x in ['考试', '成绩', '2025', '202']):
                            if score_value < 90:
                                low_score_rows.append({
                                    'row': row_idx,
                                    'score': score_value,
                                    'context': row_context
                                })
                                print(f"      行{row_idx}: 成绩 {score_value} < 90 ❌")
                    except (ValueError, TypeError):
                        pass
        
        if low_score_rows:
            for item in low_score_rows:
                file_errors.append({
                    'file': file_name,
                    'row': item['row'],
                    'check_type': '规则3',
                    'field': '考试成绩',
                    'expected': '≥90',
                    'actual': item['score'],
                    'message': f"成绩低于90，实际成绩为 {item['score']}"
                })
            print(f"      发现 {len(low_score_rows)} 条成绩低于90的记录")
        else:
            print(f"      ✅ 所有成绩均≥90")
    
    except Exception as e:
        print(f"   ❌ 文件处理异常: {e}")
        traceback.print_exc()
        file_errors.append({
            'file': file_name,
            'row': '',
            'check_type': '文件读取',
            'field': '',
            'expected': '',
            'actual': '',
            'message': f"文件读取失败: {str(e)}"
        })
    
    return file_errors


def main():
    """
    主程序
    """
    check_folder = r'd:\待检查文件夹'
    
    # 确保文件夹存在
    if not os.path.exists(check_folder):
        print(f"❌ 文件夹不存在: {check_folder}")
        os.makedirs(check_folder, exist_ok=True)
        print(f"✅ 已创建文件夹: {check_folder}")
        return
    
    # 收集所有Excel文件
    excel_files = []
    for file in os.listdir(check_folder):
        if file.endswith(('.xlsx', '.xls')) and not file.startswith('检查结果'):
            file_path = os.path.join(check_folder, file)
            excel_files.append(file_path)
    
    if not excel_files:
        print(f"⚠️ 文件夹中没有找到Excel文件: {check_folder}")
        return
    
    print(f"✅ 找到 {len(excel_files)} 个Excel文件，开始检查...\n")
    
    # 遍历每个文件进行检查
    all_errors = []
    for file_path in excel_files:
        file_errors = check_single_file(file_path)
        all_errors.extend(file_errors)
        if file_errors:
            print(f"   ⚠️ 发现 {len(file_errors)} 个错误")
        else:
            print(f"   ✅ 检查完成，无错误")
    
    # 输出结果到Excel
    output_file = os.path.join(check_folder, '检查结果.xlsx')
    
    if all_errors:
        result_df = pd.DataFrame(all_errors)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(
                writer,
                sheet_name='检查结果',
                index=False,
                columns=['file', 'row', 'check_type', 'field', 'expected', 'actual', 'message']
            )
            
            # 重命名列
            worksheet = writer.sheets['检查结果']
            for idx, col_name in enumerate(['文件名', '行号', '检查项', '字段名', '期望值', '实际值', '错误信息'], 1):
                worksheet.cell(row=1, column=idx).value = col_name
                worksheet.column_dimensions[chr(64 + idx)].width = 20
        
        print(f"\n✅ 检查完成！发现 {len(all_errors)} 个错误")
        print(f"📊 结果已保存到: {output_file}")
    else:
        result_df = pd.DataFrame({
            '文件名': ['检查完成'],
            '行号': [''],
            '检查项': [''],
            '字段名': [''],
            '期望值': [''],
            '实际值': [''],
            '错误信息': ['所有文件均检查通过，无错误']
        })
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='检查结果', index=False)
        
        print(f"\n✅ 检查完成！所有文件均检查通过，无错误")
        print(f"📊 结果已保存到: {output_file}")


if __name__ == '__main__':
    print("=" * 60)
    print("📋 表格检查工具 v2 (支持复杂非结构化表格)")
    print("=" * 60)
    
    main()
    
    print("\n✨ 程序执行完毕")
